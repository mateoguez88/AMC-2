"""
Carga datos desde el archivo Excel AMC (.xlsm/.xlsx).
Lee dinamicamente las hojas ALTERNATIVAS, INDICADORES, VALORES, Priorizacion, OBJETIVOS.
"""
from __future__ import annotations

import logging
import warnings
from dataclasses import dataclass, field

import numpy as np
import pandas as pd
import openpyxl

logger = logging.getLogger(__name__)

# Hojas requeridas en el Excel
_REQUIRED_SHEETS = {"ALTERNATIVAS", "INDICADORES", "VALORES", "Priorización", "OBJETIVOS"}


@dataclass
class AMCData:
    """Contenedor inmutable para todos los datos del modelo AMC."""

    alternatives: pd.DataFrame = field(default_factory=pd.DataFrame)
    indicators: pd.DataFrame = field(default_factory=pd.DataFrame)
    values_matrix: pd.DataFrame = field(default_factory=pd.DataFrame)
    visions: pd.DataFrame = field(default_factory=pd.DataFrame)
    objectives: pd.DataFrame = field(default_factory=pd.DataFrame)
    ahp_matrix: np.ndarray = field(default_factory=lambda: np.array([]))
    scenario: str = ""
    timeframe: str = ""

    @property
    def n_alternatives(self) -> int:
        return len(self.alternatives)

    @property
    def n_indicators(self) -> int:
        return len(self.indicators)

    @property
    def n_visions(self) -> int:
        return len(self.visions)

    @property
    def criteria_types(self) -> list[str]:
        """Lista de 'SUMA' o 'RESTA' por cada indicador, normalizada a mayusculas."""
        raw = self.indicators["SUMA O RESTA"].tolist()
        return [str(t).upper().strip() if t else "SUMA" for t in raw]

    @property
    def indicator_names(self) -> list[str]:
        return self.indicators["INDICADOR"].tolist()

    @property
    def alternative_names(self) -> list[str]:
        return self.alternatives["NOMBRE"].tolist()

    @property
    def vision_names(self) -> list[str]:
        if "VISION" in self.visions.columns:
            return self.visions["VISION"].tolist()
        return []

    @property
    def vision_weights(self) -> np.ndarray:
        if "PONDERACION" in self.visions.columns:
            return self.visions["PONDERACION"].values.astype(float)
        n = max(self.n_visions, 1)
        return np.ones(n) / n

    @property
    def vision_ids(self) -> list[str]:
        if "ID VISION" in self.visions.columns:
            return self.visions["ID VISION"].tolist()
        return []

    def validate(self) -> list[str]:
        """Valida coherencia interna de los datos. Retorna lista de warnings."""
        issues: list[str] = []
        if self.n_alternatives == 0:
            issues.append("No se encontraron alternativas.")
        if self.n_indicators == 0:
            issues.append("No se encontraron indicadores.")
        if self.values_matrix.shape[1] != self.n_indicators:
            issues.append(
                f"Desajuste: {self.values_matrix.shape[1]} columnas en valores "
                f"vs {self.n_indicators} indicadores."
            )
        if self.values_matrix.shape[0] != self.n_alternatives:
            issues.append(
                f"Desajuste: {self.values_matrix.shape[0]} filas en valores "
                f"vs {self.n_alternatives} alternativas."
            )
        # Verificar tipos de criterio validos
        for i, ct in enumerate(self.criteria_types):
            if ct not in ("SUMA", "RESTA"):
                issues.append(f"Indicador {i}: tipo '{ct}' invalido (debe ser SUMA o RESTA).")
        return issues


def _safe_headers(row: tuple, prefix: str = "col") -> list[str]:
    """Genera headers seguros a partir de una fila, evitando None y duplicados."""
    headers = []
    seen: dict[str, int] = {}
    for i, h in enumerate(row):
        name = str(h).strip() if h else f"{prefix}_{i}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        headers.append(name)
    return headers


def _find_header_row(rows: list[tuple], marker_col: str) -> int:
    """Busca dinámicamente la fila que contiene un header especifico."""
    for idx, row in enumerate(rows):
        for cell in row:
            if cell and str(cell).strip().upper() == marker_col.upper():
                return idx
    return 0


def load_from_excel(filepath: str, include_reference: bool = False) -> AMCData:
    """Carga todos los datos del archivo Excel AMC.

    Args:
        filepath: Ruta al archivo .xlsm/.xlsx.
        include_reference: Si True, incluye alternativas de referencia (Min/Max).

    Returns:
        AMCData con todos los datos cargados.

    Raises:
        ValueError: Si faltan hojas requeridas o datos criticos.
    """
    data = AMCData()

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    # Verificar hojas requeridas
    available = set(wb.sheetnames)
    missing = _REQUIRED_SHEETS - available
    if missing:
        wb.close()
        raise ValueError(f"Faltan hojas requeridas en el Excel: {missing}")

    try:
        # --- ALTERNATIVAS ---
        ws = wb["ALTERNATIVAS"]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows:
            raise ValueError("Hoja ALTERNATIVAS vacia.")
        headers = _safe_headers(rows[0])
        alt_df = pd.DataFrame(rows[1:], columns=headers)
        alt_df = alt_df.dropna(subset=["ID"])
        alt_df = alt_df[alt_df["ID"].astype(str).str.startswith("ALT")]
        alt_df = alt_df.dropna(subset=["NOMBRE"])

        if not include_reference:
            # Excluir alternativas de referencia (detectar por nombre o IDs conocidos)
            ref_ids = {"ALT00", "ALT07"}
            ref_names = {"min", "max", "referencia", "ref"}
            mask_id = alt_df["ID"].isin(ref_ids)
            mask_name = alt_df["NOMBRE"].astype(str).str.lower().str.strip().isin(ref_names)
            alt_df = alt_df[~(mask_id | mask_name)]

        data.alternatives = alt_df.reset_index(drop=True)

        # --- INDICADORES ---
        ws = wb["INDICADORES"]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        header_idx = _find_header_row(rows, "ID")
        if header_idx >= len(rows) - 1:
            raise ValueError("No se encontraron datos en hoja INDICADORES.")
        headers = _safe_headers(rows[header_idx])
        ind_df = pd.DataFrame(rows[header_idx + 1:], columns=headers)
        ind_df = ind_df.dropna(subset=["ID"])
        data.indicators = ind_df.reset_index(drop=True)

        # --- VALORES ---
        ws = wb["VALORES"]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        header_idx = _find_header_row(rows, "ID ALTERNATIVA")
        if header_idx < 0:
            # Fallback: buscar fila con contenido tipo IND
            header_idx = _find_header_row(rows, "IND01")
        headers = _safe_headers(rows[header_idx])
        # Datos empiezan despues de headers (saltar posibles sub-headers)
        data_start = header_idx + 1
        # Detectar si la siguiente fila es un sub-header (texto descriptivo)
        if data_start < len(rows):
            first_cell = rows[data_start][0]
            if first_cell and not str(first_cell).startswith("ALT"):
                data_start += 1

        val_df = pd.DataFrame(rows[data_start:], columns=headers)
        id_col = headers[0]
        val_df = val_df.dropna(subset=[id_col])

        if not include_reference:
            ref_ids = {"ALT00", "ALT07"}
            val_df = val_df[~val_df[id_col].astype(str).isin(ref_ids)]

        alt_ids = val_df[id_col].tolist()
        indicator_cols = [c for c in headers[1:] if str(c).startswith("IND")]

        values_matrix = val_df[indicator_cols].apply(pd.to_numeric, errors="coerce")

        # Registrar NaN antes de rellenar
        nan_count = values_matrix.isna().sum().sum()
        if nan_count > 0:
            warnings.warn(
                f"Se encontraron {nan_count} valores vacios/no numericos en la matriz. "
                f"Se reemplazaron por 0.",
                UserWarning,
                stacklevel=2,
            )
            logger.warning("Valores NaN encontrados en matriz de valores: %d", nan_count)
        values_matrix = values_matrix.fillna(0)
        values_matrix.index = alt_ids
        data.values_matrix = values_matrix

        # --- PRIORIZACION (Visiones y AHP) ---
        ws = wb["Priorización"]
        all_rows = list(ws.iter_rows(values_only=True))

        # Escenario (fila 3, index 2)
        if len(all_rows) >= 3:
            data.scenario = str(all_rows[2][0]) if all_rows[2][0] else ""
            data.timeframe = str(all_rows[2][1]) if len(all_rows[2]) > 1 and all_rows[2][1] else ""

        # Buscar header de visiones dinámicamente
        vision_header_idx = _find_header_row(all_rows, "ID VISION")
        if vision_header_idx < 0:
            vision_header_idx = 5  # Fallback al indice original

        vision_headers = _safe_headers(all_rows[vision_header_idx])
        vision_data = []
        for r in all_rows[vision_header_idx + 1: vision_header_idx + 20]:
            if r and r[0] is not None and str(r[0]).startswith("V"):
                vision_data.append(r)

        vision_df = pd.DataFrame(vision_data, columns=vision_headers)
        data.visions = vision_df

        # Extraer matriz AHP
        ahp_col_start = None
        for i, h in enumerate(vision_headers):
            if str(h).strip() == "V1":
                ahp_col_start = i
                break

        if ahp_col_start is not None:
            n_visions = len(vision_data)
            ahp_matrix = np.ones((n_visions, n_visions))
            for i, row in enumerate(vision_data):
                for j in range(n_visions):
                    col_idx = ahp_col_start + j
                    if col_idx < len(row):
                        val = row[col_idx]
                        if val is not None:
                            try:
                                ahp_matrix[i, j] = float(val)
                            except (ValueError, TypeError):
                                logger.warning("Valor AHP invalido en [%d,%d]: %s", i, j, val)
            data.ahp_matrix = ahp_matrix

        # --- OBJETIVOS ---
        ws = wb["OBJETIVOS"]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        header_idx = _find_header_row(rows, "ID OBJETIVO")
        if header_idx < 0:
            header_idx = 5
        headers = _safe_headers(rows[header_idx])
        obj_df = pd.DataFrame(rows[header_idx + 1:], columns=headers)
        obj_df = obj_df.dropna(subset=["ID OBJETIVO"])
        data.objectives = obj_df.reset_index(drop=True)

    finally:
        wb.close()

    # Validar datos cargados
    issues = data.validate()
    for issue in issues:
        logger.warning("Validacion: %s", issue)
        warnings.warn(f"AMC Data: {issue}", UserWarning, stacklevel=2)

    return data


def get_decision_matrix(data: AMCData) -> pd.DataFrame:
    """Obtiene la matriz de decision lista para analisis.

    Filas: alternativas (con nombres como indice).
    Columnas: indicadores.
    """
    dm = data.values_matrix.copy()
    alt_name_map = dict(zip(data.alternatives["ID"], data.alternatives["NOMBRE"]))
    dm.index = [alt_name_map.get(aid, aid) for aid in dm.index]
    return dm


def get_weights_from_data(data: AMCData) -> np.ndarray:
    """Obtiene los pesos ponderados de cada indicador desde los datos cargados."""
    if "PESO_PONDERADO" not in data.indicators.columns:
        logger.warning("Columna PESO_PONDERADO no encontrada, usando pesos iguales.")
        n = data.n_indicators or 1
        return np.ones(n) / n
    weights = data.indicators["PESO_PONDERADO"].values.astype(float)
    total = weights.sum()
    if total > 0:
        weights = weights / total
    return weights


def get_linear_weights(data: AMCData) -> np.ndarray:
    """Obtiene los pesos lineales (iguales) de cada indicador."""
    if "PESO_LINEAL" not in data.indicators.columns:
        n = data.n_indicators or 1
        return np.ones(n) / n
    weights = data.indicators["PESO_LINEAL"].values.astype(float)
    total = weights.sum()
    if total > 0:
        weights = weights / total
    return weights


def get_hierarchy_mapping(data: AMCData) -> dict:
    """Obtiene el mapeo jerarquico Vision > Objetivo > Indicadores."""
    hierarchy: dict = {}
    for _, ind in data.indicators.iterrows():
        v_id = ind.get("ID VISION", "")
        o_id = ind.get("ID OBJETIVO", "")
        i_id = ind.get("ID", "")

        if v_id not in hierarchy:
            hierarchy[v_id] = {"objectives": {}}
        if o_id not in hierarchy[v_id]["objectives"]:
            hierarchy[v_id]["objectives"][o_id] = []
        hierarchy[v_id]["objectives"][o_id].append(i_id)

    return hierarchy


def group_by_vision(
    normalized: pd.DataFrame,
    indicators: pd.DataFrame,
) -> pd.DataFrame:
    """Agrupa columnas de indicadores por vision, promediando valores.

    Util para radar charts, heatmaps agrupados, etc.
    """
    indicator_vision = indicators["ID VISION"].tolist()
    vision_ids = sorted(set(indicator_vision))

    grouped = pd.DataFrame(index=normalized.index)
    for v_id in vision_ids:
        cols = [c for c, v in zip(normalized.columns, indicator_vision) if v == v_id]
        if cols:
            grouped[v_id] = normalized[cols].mean(axis=1)
    return grouped
