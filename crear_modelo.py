"""
Script para generar un Excel modelo AMC con inputs clave editables.
Ejecutar: python crear_modelo.py
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, CellIsRule
import os

OUTPUT = os.path.join(os.path.dirname(__file__), "AMC_Modelo.xlsx")

# ── Estilos ──────────────────────────────────────────────────────────
BLUE = "1F77B4"
DARK = "2C3E50"
GREEN = "27AE60"
ORANGE = "E67E22"
RED = "E74C3C"
PURPLE = "8E44AD"
LIGHT_BLUE = "D6EAF8"
LIGHT_GREEN = "D5F5E3"
LIGHT_ORANGE = "FDEBD0"
LIGHT_GREY = "F2F3F4"
WHITE = "FFFFFF"

header_font = Font(bold=True, color=WHITE, size=11)
header_fill = PatternFill("solid", fgColor=BLUE)
section_font = Font(bold=True, color=DARK, size=13)
subsection_font = Font(bold=True, color=BLUE, size=11)
input_fill = PatternFill("solid", fgColor=LIGHT_GREEN)
locked_fill = PatternFill("solid", fgColor=LIGHT_GREY)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header_row(ws, row, cols, fill=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = fill or header_fill
        cell.alignment = center
        cell.border = thin_border


def style_data_area(ws, start_row, end_row, cols, editable_cols=None):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = center
            if editable_cols and c in editable_cols:
                cell.fill = input_fill
            elif editable_cols:
                cell.fill = locked_fill


def auto_width(ws, min_width=12, max_width=30):
    for col in ws.columns:
        length = min_width
        for cell in col:
            if cell.value:
                length = max(length, min(len(str(cell.value)) + 4, max_width))
        ws.column_dimensions[get_column_letter(col[0].column)].width = length


# ── Datos de ejemplo ─────────────────────────────────────────────────
VISIONES = [
    ("V1", "Ambiental", 0.20),
    ("V2", "Social", 0.15),
    ("V3", "Economica", 0.20),
    ("V4", "Tecnica", 0.15),
    ("V5", "Territorial", 0.10),
    ("V6", "Gobernanza", 0.08),
    ("V7", "Riesgo", 0.07),
    ("V8", "Innovacion", 0.05),
]

OBJETIVOS = [
    ("OBJ01", "V1", "Minimizar impacto ambiental"),
    ("OBJ02", "V1", "Proteger biodiversidad"),
    ("OBJ03", "V2", "Maximizar beneficio social"),
    ("OBJ04", "V3", "Optimizar costos"),
    ("OBJ05", "V3", "Maximizar retorno economico"),
    ("OBJ06", "V4", "Viabilidad tecnica"),
    ("OBJ07", "V5", "Integracion territorial"),
    ("OBJ08", "V6", "Gobernanza participativa"),
    ("OBJ09", "V7", "Minimizar riesgos"),
]

INDICADORES = [
    # (ID, ID_VISION, ID_OBJ, INDICADOR, UNIDAD, SUMA/RESTA, PESO_LINEAL, PESO_PONDERADO)
    ("IND01", "V1", "OBJ01", "Emisiones CO2 evitadas", "tCO2/año", "SUMA", 1, 0.06),
    ("IND02", "V1", "OBJ01", "Consumo energetico", "kWh/año", "RESTA", 1, 0.05),
    ("IND03", "V1", "OBJ02", "Area natural protegida", "ha", "SUMA", 1, 0.05),
    ("IND04", "V1", "OBJ02", "Indice biodiversidad", "adim.", "SUMA", 1, 0.04),
    ("IND05", "V2", "OBJ03", "Empleos generados", "empleos", "SUMA", 1, 0.05),
    ("IND06", "V2", "OBJ03", "Poblacion beneficiada", "hab.", "SUMA", 1, 0.05),
    ("IND07", "V2", "OBJ03", "Indice equidad", "adim.", "SUMA", 1, 0.05),
    ("IND08", "V3", "OBJ04", "Costo inversion", "M€", "RESTA", 1, 0.06),
    ("IND09", "V3", "OBJ04", "Costo operacion anual", "M€/año", "RESTA", 1, 0.05),
    ("IND10", "V3", "OBJ05", "VAN", "M€", "SUMA", 1, 0.05),
    ("IND11", "V3", "OBJ05", "TIR", "%", "SUMA", 1, 0.04),
    ("IND12", "V4", "OBJ06", "Madurez tecnologica (TRL)", "nivel", "SUMA", 1, 0.05),
    ("IND13", "V4", "OBJ06", "Plazo implementacion", "meses", "RESTA", 1, 0.05),
    ("IND14", "V4", "OBJ06", "Disponibilidad recursos", "%", "SUMA", 1, 0.05),
    ("IND15", "V5", "OBJ07", "Compatibilidad territorial", "adim.", "SUMA", 1, 0.05),
    ("IND16", "V5", "OBJ07", "Accesibilidad", "adim.", "SUMA", 1, 0.05),
    ("IND17", "V6", "OBJ08", "Apoyo institucional", "adim.", "SUMA", 1, 0.04),
    ("IND18", "V6", "OBJ08", "Participacion ciudadana", "adim.", "SUMA", 1, 0.04),
    ("IND19", "V7", "OBJ09", "Riesgo tecnico", "adim.", "RESTA", 1, 0.04),
    ("IND20", "V7", "OBJ09", "Riesgo financiero", "adim.", "RESTA", 1, 0.03),
]

ALTERNATIVAS = [
    ("ALT01", "Alternativa A - Base"),
    ("ALT02", "Alternativa B - Moderada"),
    ("ALT03", "Alternativa C - Ambiciosa"),
    ("ALT04", "Alternativa D - Sostenible"),
    ("ALT05", "Alternativa E - Economica"),
]

# Valores de ejemplo (cada fila = alternativa, cada col = indicador)
VALORES = [
    # ALT01: Base
    [5000, 120000, 50, 0.6, 200, 15000, 0.5, 80, 5.0, 20, 8, 7, 24, 70, 0.6, 0.7, 0.5, 0.4, 0.3, 0.4],
    # ALT02: Moderada
    [8000, 100000, 80, 0.7, 350, 25000, 0.6, 120, 7.0, 35, 12, 8, 18, 80, 0.7, 0.8, 0.6, 0.5, 0.4, 0.3],
    # ALT03: Ambiciosa
    [15000, 80000, 120, 0.85, 500, 40000, 0.8, 200, 12.0, 60, 18, 6, 36, 60, 0.8, 0.6, 0.7, 0.7, 0.6, 0.5],
    # ALT04: Sostenible
    [12000, 90000, 150, 0.9, 400, 35000, 0.9, 150, 8.0, 45, 15, 7, 30, 75, 0.9, 0.8, 0.8, 0.8, 0.3, 0.2],
    # ALT05: Economica
    [3000, 150000, 30, 0.4, 150, 10000, 0.4, 50, 3.0, 15, 6, 9, 12, 90, 0.5, 0.9, 0.4, 0.3, 0.2, 0.2],
]


def create_workbook():
    wb = openpyxl.Workbook()

    # ================================================================
    # HOJA: INSTRUCCIONES
    # ================================================================
    ws = wb.active
    ws.title = "INSTRUCCIONES"
    ws.sheet_properties.tabColor = BLUE

    instructions = [
        ("AMC - MODELO DE ANALISIS MULTICRITERIO", section_font),
        ("", None),
        ("Este archivo contiene los inputs clave para ejecutar el analisis multicriterio.", None),
        ("Las celdas en VERDE son editables. Las celdas en GRIS son calculadas o de referencia.", None),
        ("", None),
        ("HOJAS DEL MODELO:", subsection_font),
        ("  1. ALTERNATIVAS  → Definir las alternativas a evaluar", None),
        ("  2. INDICADORES   → Definir indicadores, sus unidades, tipo y pesos", None),
        ("  3. VALORES       → Ingresar los valores de cada alternativa por indicador", None),
        ("  4. VISIONES      → Definir visiones y sus ponderaciones", None),
        ("  5. OBJETIVOS     → Definir objetivos por vision", None),
        ("  6. AHP           → Matriz de comparacion pareada AHP (opcional)", None),
        ("", None),
        ("FLUJO DE TRABAJO:", subsection_font),
        ("  1. Completar ALTERNATIVAS con los nombres de sus alternativas", None),
        ("  2. Ajustar INDICADORES: agregar/quitar, definir tipo SUMA/RESTA y pesos", None),
        ("  3. Llenar VALORES con los datos reales de cada alternativa", None),
        ("  4. Ajustar pesos de VISIONES (deben sumar 1.0)", None),
        ("  5. (Opcional) Completar la matriz AHP para pesos por comparacion pareada", None),
        ("  6. Cargar este archivo en la app Streamlit: streamlit run app/main.py", None),
        ("", None),
        ("CONVENCIONES:", subsection_font),
        ("  SUMA  = indicador de beneficio (mayor es mejor)", None),
        ("  RESTA = indicador de costo (menor es mejor)", None),
        ("  Los pesos de indicadores deben sumar 1.0 (se normalizan automaticamente)", None),
    ]

    for i, (text, font) in enumerate(instructions, 1):
        cell = ws.cell(row=i, column=1, value=text)
        if font:
            cell.font = font
    ws.column_dimensions["A"].width = 80

    # ================================================================
    # HOJA: ALTERNATIVAS
    # ================================================================
    ws = wb.create_sheet("ALTERNATIVAS")
    ws.sheet_properties.tabColor = GREEN

    headers = ["ID", "NOMBRE"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for i, (aid, name) in enumerate(ALTERNATIVAS, 2):
        ws.cell(row=i, column=1, value=aid).fill = locked_fill
        ws.cell(row=i, column=2, value=name).fill = input_fill

    style_data_area(ws, 2, len(ALTERNATIVAS) + 1, 2, editable_cols={2})

    # Nota para agregar mas alternativas
    note_row = len(ALTERNATIVAS) + 3
    ws.cell(row=note_row, column=1, value="Para agregar mas alternativas:").font = Font(italic=True, color="666666")
    ws.cell(row=note_row + 1, column=1, value="Insertar filas con ID (ALT06, ALT07...) y nombre").font = Font(italic=True, color="666666")
    auto_width(ws)

    # ================================================================
    # HOJA: INDICADORES
    # ================================================================
    ws = wb.create_sheet("INDICADORES")
    ws.sheet_properties.tabColor = ORANGE

    # 4 filas de encabezado de contexto (data_loader lee desde fila 5)
    ws.cell(row=1, column=1, value="TABLA DE INDICADORES").font = section_font
    ws.cell(row=2, column=1, value="Definir los indicadores del modelo AMC")
    ws.cell(row=3, column=1, value="Celdas VERDES = editables")
    ws.cell(row=4, column=1)

    headers = [
        "ID", "ID VISION", "ID OBJETIVO", "INDICADOR", "UNIDAD",
        "SUMA O RESTA", "PESO_LINEAL", "PESO_PONDERADO"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=5, column=c, value=h)
    style_header_row(ws, 5, len(headers))

    editable = {4, 5, 6, 7, 8}  # INDICADOR, UNIDAD, SUMA/RESTA, pesos
    for i, ind in enumerate(INDICADORES):
        row = 6 + i
        for c, val in enumerate(ind, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.border = thin_border
            cell.alignment = center
            if c in editable:
                cell.fill = input_fill
            else:
                cell.fill = locked_fill

    # Validacion SUMA/RESTA
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"SUMA,RESTA"', allow_blank=False)
    dv.error = "Solo se permite SUMA o RESTA"
    dv.errorTitle = "Valor invalido"
    ws.add_data_validation(dv)
    for r in range(6, 6 + len(INDICADORES)):
        dv.add(ws.cell(row=r, column=6))

    auto_width(ws, min_width=14)

    # ================================================================
    # HOJA: VALORES
    # ================================================================
    ws = wb.create_sheet("VALORES")
    ws.sheet_properties.tabColor = RED

    ws.cell(row=1, column=1, value="MATRIZ DE VALORES").font = section_font
    ws.cell(row=2, column=1, value="Ingresar valores de cada alternativa por indicador (celdas VERDES)")

    # Headers fila 3
    val_headers = ["ID ALTERNATIVA"] + [ind[0] for ind in INDICADORES]
    for c, h in enumerate(val_headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, len(val_headers))

    # Subheader con nombres de indicador
    for c, ind in enumerate(INDICADORES, 2):
        cell = ws.cell(row=4, column=c, value=ind[3])  # nombre del indicador
        cell.font = Font(italic=True, size=8, color="666666")
        cell.alignment = center
        cell.border = thin_border

    # Datos
    for i, (alt_id, _) in enumerate(ALTERNATIVAS):
        row = 5 + i
        ws.cell(row=row, column=1, value=alt_id).fill = locked_fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = center
        for c, val in enumerate(VALORES[i], 2):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill = input_fill
            cell.border = thin_border
            cell.alignment = center
            if isinstance(val, float) and val < 1:
                cell.number_format = "0.00"

    # Formato condicional (data bars) en cada columna de indicadores
    for c in range(2, len(INDICADORES) + 2):
        col_letter = get_column_letter(c)
        rule = DataBarRule(
            start_type="min", end_type="max",
            color="5B9BD5", showValue=True,
        )
        ws.conditional_formatting.add(
            f"{col_letter}5:{col_letter}{4 + len(ALTERNATIVAS)}", rule
        )

    auto_width(ws, min_width=14)

    # ================================================================
    # HOJA: VISIONES (Priorización)
    # ================================================================
    ws = wb.create_sheet("Priorización")
    ws.sheet_properties.tabColor = PURPLE

    ws.cell(row=1, column=1, value="PRIORIZACION DE VISIONES").font = section_font
    ws.cell(row=2, column=1, value="Configuracion del escenario:")
    # Row 3 = index 2 in data_loader: scenario in col A, timeframe in col B
    ws.cell(row=3, column=1, value="Base").fill = input_fill
    ws.cell(row=3, column=2, value="Medio plazo").fill = input_fill
    ws.cell(row=3, column=3, value="← Escenario / Plazo (editables)").font = Font(italic=True, color="666666")
    ws.cell(row=4, column=1)
    ws.cell(row=5, column=1, value="Ajustar las ponderaciones (deben sumar 1.0)")

    # Escenario y timeframe en fila 3 (indice 2 para data_loader)
    # data_loader espera: all_rows[2][0] = scenario, all_rows[2][1] = timeframe

    # Headers fila 6
    v_headers = [
        "ID VISION", "VISION", "DESCRIPCION", "PONDERACION",
        "PONDERACION LINEAL", "PONDERACION PONDERADA", "TIPO PONDERACION",
        "V1", "V2", "V3", "V4", "V5", "V6", "V7", "V8"
    ]
    for c, h in enumerate(v_headers, 1):
        ws.cell(row=6, column=c, value=h)
    style_header_row(ws, 6, len(v_headers))

    vision_descriptions = [
        "Impacto ambiental y sostenibilidad",
        "Beneficios sociales y equidad",
        "Viabilidad y retorno economico",
        "Capacidad tecnica y tecnologica",
        "Integracion con el territorio",
        "Participacion y gobernanza",
        "Gestion de riesgos",
        "Innovacion y desarrollo",
    ]

    # Matriz AHP identidad (para que el usuario la complete)
    ahp_default = [
        [1, 2, 1, 2, 3, 3, 3, 4],
        [0.5, 1, 0.5, 1, 2, 2, 2, 3],
        [1, 2, 1, 2, 3, 3, 3, 4],
        [0.5, 1, 0.5, 1, 2, 2, 2, 3],
        [0.33, 0.5, 0.33, 0.5, 1, 1, 1, 2],
        [0.33, 0.5, 0.33, 0.5, 1, 1, 1, 2],
        [0.33, 0.5, 0.33, 0.5, 1, 1, 1, 2],
        [0.25, 0.33, 0.25, 0.33, 0.5, 0.5, 0.5, 1],
    ]

    for i, (vid, vname, pond) in enumerate(VISIONES):
        row = 7 + i
        ws.cell(row=row, column=1, value=vid).fill = locked_fill
        ws.cell(row=row, column=2, value=vname).fill = input_fill
        ws.cell(row=row, column=3, value=vision_descriptions[i]).fill = input_fill
        ws.cell(row=row, column=4, value=pond).fill = input_fill
        ws.cell(row=row, column=4).number_format = "0.00"
        ws.cell(row=row, column=5, value=1.0 / len(VISIONES)).fill = locked_fill
        ws.cell(row=row, column=5).number_format = "0.0000"
        ws.cell(row=row, column=6, value=pond).fill = locked_fill
        ws.cell(row=row, column=6).number_format = "0.0000"
        ws.cell(row=row, column=7, value="PONDERADA").fill = locked_fill

        # Matriz AHP
        for j in range(8):
            cell = ws.cell(row=row, column=8 + j, value=ahp_default[i][j])
            cell.fill = input_fill
            cell.border = thin_border
            cell.alignment = center
            cell.number_format = "0.00"

    style_data_area(ws, 7, 7 + len(VISIONES) - 1, len(v_headers),
                    editable_cols={2, 3, 4, 8, 9, 10, 11, 12, 13, 14, 15})

    # Suma de pesos
    sum_row = 7 + len(VISIONES)
    ws.cell(row=sum_row, column=3, value="SUMA:").font = Font(bold=True)
    ws.cell(row=sum_row, column=4).value = f"=SUM(D7:D{sum_row-1})"
    ws.cell(row=sum_row, column=4).font = Font(bold=True, color=RED)
    ws.cell(row=sum_row, column=4).number_format = "0.00"

    # Formato condicional: si la suma != 1, rojo
    ws.conditional_formatting.add(
        f"D{sum_row}",
        CellIsRule(operator="notEqual", formula=["1"], fill=PatternFill("solid", fgColor="FFCCCC"))
    )
    ws.conditional_formatting.add(
        f"D{sum_row}",
        CellIsRule(operator="equal", formula=["1"], fill=PatternFill("solid", fgColor="CCFFCC"))
    )

    auto_width(ws, min_width=12)

    # ================================================================
    # HOJA: OBJETIVOS
    # ================================================================
    ws = wb.create_sheet("OBJETIVOS")
    ws.sheet_properties.tabColor = "3498DB"

    ws.cell(row=1, column=1, value="OBJETIVOS POR VISION").font = section_font
    for r in range(2, 6):
        ws.cell(row=r, column=1)

    headers = ["ID OBJETIVO", "ID VISION", "OBJETIVO", "DESCRIPCION"]
    ws.cell(row=6, column=1, value="ID OBJETIVO")
    ws.cell(row=6, column=2, value="ID VISION")
    ws.cell(row=6, column=3, value="OBJETIVO")
    ws.cell(row=6, column=4, value="DESCRIPCION")
    style_header_row(ws, 6, 4)

    obj_descriptions = [
        "Reducir emisiones y consumo de recursos",
        "Conservar ecosistemas y especies",
        "Generar empleo y bienestar comunitario",
        "Minimizar costos de inversion y operacion",
        "Maximizar retorno financiero",
        "Asegurar factibilidad tecnica",
        "Minimizar impacto territorial",
        "Fomentar transparencia y participacion",
        "Identificar y mitigar riesgos",
    ]

    for i, (oid, vid, oname) in enumerate(OBJETIVOS):
        row = 7 + i
        ws.cell(row=row, column=1, value=oid).fill = locked_fill
        ws.cell(row=row, column=2, value=vid).fill = locked_fill
        ws.cell(row=row, column=3, value=oname).fill = input_fill
        ws.cell(row=row, column=4, value=obj_descriptions[i]).fill = input_fill

    style_data_area(ws, 7, 7 + len(OBJETIVOS) - 1, 4, editable_cols={3, 4})
    auto_width(ws)

    # ================================================================
    # HOJA: RESUMEN (calculado)
    # ================================================================
    ws = wb.create_sheet("RESUMEN")
    ws.sheet_properties.tabColor = "2ECC71"

    ws.cell(row=1, column=1, value="RESUMEN DEL MODELO").font = section_font
    ws.cell(row=2, column=1)

    summary_data = [
        ("Parametro", "Valor"),
        ("Numero de alternativas", len(ALTERNATIVAS)),
        ("Numero de indicadores", len(INDICADORES)),
        ("Numero de visiones", len(VISIONES)),
        ("Numero de objetivos", len(OBJETIVOS)),
        ("Indicadores beneficio (SUMA)", sum(1 for i in INDICADORES if i[5] == "SUMA")),
        ("Indicadores costo (RESTA)", sum(1 for i in INDICADORES if i[5] == "RESTA")),
    ]

    for r, (param, val) in enumerate(summary_data, 3):
        ws.cell(row=r, column=1, value=param)
        ws.cell(row=r, column=2, value=val)
        if r == 3:
            ws.cell(row=r, column=1).font = header_font
            ws.cell(row=r, column=1).fill = header_fill
            ws.cell(row=r, column=2).font = header_font
            ws.cell(row=r, column=2).fill = header_fill
        else:
            ws.cell(row=r, column=1).fill = locked_fill
            ws.cell(row=r, column=2).fill = locked_fill
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2).border = thin_border

    # Tabla de pesos por vision
    ws.cell(row=12, column=1, value="DISTRIBUCION DE PESOS POR VISION").font = subsection_font
    ws.cell(row=13, column=1, value="Vision").font = header_font
    ws.cell(row=13, column=1).fill = header_fill
    ws.cell(row=13, column=2, value="Peso").font = header_font
    ws.cell(row=13, column=2).fill = header_fill
    ws.cell(row=13, column=3, value="% Total").font = header_font
    ws.cell(row=13, column=3).fill = header_fill

    for i, (vid, vname, pond) in enumerate(VISIONES):
        row = 14 + i
        ws.cell(row=row, column=1, value=f"{vid} - {vname}").border = thin_border
        ws.cell(row=row, column=2, value=pond).border = thin_border
        ws.cell(row=row, column=2).number_format = "0.00"
        ws.cell(row=row, column=3, value=pond * 100).border = thin_border
        ws.cell(row=row, column=3).number_format = "0.0\"%\""

    auto_width(ws, min_width=20)
    ws.column_dimensions["A"].width = 35

    # ================================================================
    # Guardar
    # ================================================================
    wb.save(OUTPUT)
    print(f"Excel modelo generado: {OUTPUT}")
    return OUTPUT


if __name__ == "__main__":
    create_workbook()
